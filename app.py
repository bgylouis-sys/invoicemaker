from flask import Flask, render_template, request, send_file, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from copy import copy
import io, os, re, json, uuid, urllib.request, urllib.parse, base64
from datetime import date, datetime
from PIL import Image as PILImage
from excel_utils import format_invoice_date, shift_item_rows, fill_bank_from_json, insert_image as excel_insert_image

app = Flask(__name__)

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH   = os.path.join(BASE_DIR, '约旦卫浴Proforma Invoice模板.xlsx')
LOGO_DIR        = os.path.join(BASE_DIR, 'logo')
LOGO_NAMES_PATH = os.path.join(LOGO_DIR, 'names.json')
CUSTOMERS_PATH  = os.path.join(BASE_DIR, 'customers.json')
PRODUCTS_PATH   = os.path.join(BASE_DIR, 'products.json')
BANK_INFO_PATH  = os.path.join(BASE_DIR, 'bank_info.json')
LOGO_EXTS       = {'.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'}

ITEMS_START = 17
ITEMS_END   = 21
ITEMS_SLOTS = 5

CERAMIC_PI_TEMPLATE_PATH = os.path.join(BASE_DIR, '约旦瓷砖PI模板.xlsx')

TEMPLATE_PATHS = {
    'default': TEMPLATE_PATH,
    'LED灯具': os.path.join(BASE_DIR, '约旦LED Proforma Invoice模板.xlsx'),
}


# ── Logo helpers ──────────────────────────────────────────────────────────────

def list_logo_files():
    return sorted(
        f for f in os.listdir(LOGO_DIR)
        if os.path.splitext(f)[1].lower() in LOGO_EXTS
    )

def load_logo_names():
    if os.path.exists(LOGO_NAMES_PATH):
        with open(LOGO_NAMES_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_logo_names(names):
    with open(LOGO_NAMES_PATH, 'w', encoding='utf-8') as f:
        json.dump(names, f, ensure_ascii=False, indent=2)

def logo_display_name(filename, names):
    return names.get(filename, os.path.splitext(filename)[0].upper())

def logo_list_with_names():
    files = list_logo_files()
    names = load_logo_names()
    return [{'filename': f, 'display_name': logo_display_name(f, names)} for f in files]


# ── Customer helpers ──────────────────────────────────────────────────────────

def load_customers():
    if os.path.exists(CUSTOMERS_PATH):
        with open(CUSTOMERS_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_customers_file(customers):
    with open(CUSTOMERS_PATH, 'w', encoding='utf-8') as f:
        json.dump(customers, f, ensure_ascii=False, indent=2)


# ── Product helpers ───────────────────────────────────────────────────────────

def load_products():
    if os.path.exists(PRODUCTS_PATH):
        with open(PRODUCTS_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_products_file(products):
    with open(PRODUCTS_PATH, 'w', encoding='utf-8') as f:
        json.dump(products, f, ensure_ascii=False, indent=2)


# ── Bank info helpers ────────────────────────────────────────────────────────

def load_bank_info():
    if os.path.exists(BANK_INFO_PATH):
        with open(BANK_INFO_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_bank_info_file(data):
    with open(BANK_INFO_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── Excel helpers ─────────────────────────────────────────────────────────────

def copy_row_style(ws, src_row, dst_row, max_col=14):
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font      = copy(src.font)
            dst.fill      = copy(src.fill)
            dst.border    = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


def insert_image(ws, file_obj, anchor, w=100, h=75):
    try:
        raw = file_obj.read()
        pil = PILImage.open(io.BytesIO(raw))
        pil.thumbnail((w, h), PILImage.LANCZOS)
        buf = io.BytesIO()
        fmt = (pil.format or 'PNG').upper()
        if fmt not in ('PNG', 'JPEG', 'GIF', 'BMP'):
            fmt = 'PNG'
        if fmt == 'JPEG' and pil.mode == 'RGBA':
            pil = pil.convert('RGB')
        pil.save(buf, format=fmt)
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.anchor = anchor
        ws.add_image(xl_img)
    except Exception as e:
        print(f'[image] {anchor}: {e}')


def fill_template(form, files, lang='zh', mode='pi'):
    project = form.get('project', '').strip()
    if project == '瓷砖':
        return fill_ceramic_pi(form, files, lang, mode=mode)

    template_path = TEMPLATE_PATHS.get(project, TEMPLATE_PATHS['default'])
    with open(template_path, 'rb') as f:
        buf = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    is_ar = (lang == 'ar')
    n = max(1, int(form.get('item_count') or 1))

    # ── Supplier info (rows 4-8 left, value col=C) ───────────────────────────
    ws.cell(4, 3).value = form.get('company_name', '')
    ws.cell(5, 3).value = form.get('address', '')
    ws.cell(6, 3).value = form.get('city_country', '')
    tel = (form.get('tel') or '').strip()
    ws.cell(7, 3).value = tel if tel else None
    ws.cell(8, 3).value = form.get('email', '')

    # ── Invoice info + key terms (rows 4-8 right, value col=J) ──────────────
    ws.cell(4, 10).value = form.get('invoice_no', '')
    date_raw = form.get('date', '')
    if date_raw and '-' in date_raw:
        try:
            d_obj = datetime.strptime(date_raw, '%Y-%m-%d')
            date_str = f"{d_obj.year}.{d_obj.month}.{d_obj.day}"
        except ValueError:
            date_str = date_raw
    else:
        date_str = date_raw
    ws.cell(5, 10).value = date_str
    currency = form.get('currency') or 'JOD'
    ws.cell(6, 10).value = currency
    ws.cell(7, 10).value = form.get('incoterms', '')
    ws.cell(8, 10).value = form.get('payment_terms', '')

    # ── Customer info (rows 10-14 left, value col=C) ──────────────────────────
    ws.cell(10, 3).value = form.get('customer_name', '')
    ws.cell(11, 3).value = form.get('customer_address', '')
    ws.cell(12, 3).value = form.get('customer_email', '')
    ws.cell(13, 3).value = form.get('country', '')
    cust_tel = (form.get('customer_tel') or '').strip()
    ws.cell(14, 3).value = cust_tel if cust_tel else None

    # ── Quotation details (rows 10-14 right, value col=J) ──────────────────────
    prefix = (form.get('contact_prefix') or '').strip()
    person = (form.get('contact_person') or '').strip()
    ws.cell(10, 10).value = f"{prefix} {person}".strip()
    ws.cell(11, 10).value = form.get('inquiry_ref', '')
    ws.cell(12, 10).value = form.get('lead_time', '')
    validity = (form.get('quotation_validity') or '').strip()
    ws.cell(13, 10).value = f'{validity} days' if validity else ''
    port = (form.get('port_of_loading') or '').strip()
    dest = (form.get('destination') or '').strip()
    if port and dest:    shipping = f'{port} → {dest}'
    elif port:          shipping = port
    else:               shipping = dest
    ws.cell(14, 10).value = shipping

    # ── Overwrite section headers for Arabic ──────────────────────────────────
    if is_ar:
        ws.cell(3,  1).value = 'FROM / SUPPLIER  المورّد'
        ws.cell(3,  7).value = 'INVOICE INFO  معلومات الفاتورة'
        ws.cell(9,  1).value = 'TO / CUSTOMER  العميل'
        ws.cell(9,  7).value = 'QUOTATION DETAILS  تفاصيل العرض'
        ws.cell(15, 1).value = 'ITEM LIST  قائمة المنتجات'
        ws.cell(25, 1).value = 'REMARKS  ملاحظات'
        ws.cell(31, 1).value = 'TRANSFER DETAILS  معلومات التحويل'
        ws.cell(36, 1).value = 'Company Stamp  ختم الشركة:'

    # ── CI title override ──────────────────────────────────────────────────
    if mode == 'ci':
        title_cell = ws.cell(1, 7)
        if title_cell.value and 'PROFORMA' in str(title_cell.value):
            title_cell.value = str(title_cell.value).replace('PROFORMA INVOICE', 'COMMERCIAL INVOICE')

    # ── Insert / delete item rows ─────────────────────────────────────────
    # openpyxl's insert_rows / delete_rows corrupt merged-cell references,
    # so we save all merges below the affected area, remove them, perform the
    # row operation, then re-merge at the correct shifted positions.
    offset = 0
    deleted = 0

    # Helper: collect merged ranges at or below a given row
    def _merges_below(ws, start_row):
        out = []
        for rng in ws.merged_cells.ranges:
            if rng.min_row >= start_row:
                out.append((rng.min_row, rng.max_row, rng.min_col, rng.max_col))
        return out

    # Helper: remove merged ranges at or below a given row (direct list removal,
    # no _cells touch — avoids KeyError on orphaned references)
    def _remove_merges_below(ws, start_row):
        to_del = [rng for rng in ws.merged_cells.ranges if rng.min_row >= start_row]
        for rng in to_del:
            ws.merged_cells.ranges.remove(rng)

    # Helper: re-merge saved ranges after applying a row shift
    def _reapply_merges(ws, saved, row_shift):
        for min_r, max_r, min_c, max_c in saved:
            nr_min = min_r + row_shift
            nr_max = max_r + row_shift
            if nr_min >= ITEMS_START:
                ws.merge_cells(start_row=nr_min, start_column=min_c,
                               end_row=nr_max, end_column=max_c)

    if n > ITEMS_SLOTS:
        extra = n - ITEMS_SLOTS
        offset = extra
        insert_at = ITEMS_END + 1

        saved = _merges_below(ws, insert_at)
        _remove_merges_below(ws, insert_at)

        ws.insert_rows(insert_at, extra)

        # Remove any phantom ranges left at the insertion point
        stale = [rng for rng in ws.merged_cells.ranges
                 if insert_at <= rng.min_row < insert_at + extra]
        for rng in stale:
            ws.merged_cells.ranges.remove(rng)

        _reapply_merges(ws, saved, extra)

        for i in range(extra):
            copy_row_style(ws, ITEMS_END, insert_at + i)

    if n < ITEMS_SLOTS:
        deleted = ITEMS_SLOTS - n
        delete_at = ITEMS_START + n

        saved = _merges_below(ws, delete_at)
        _remove_merges_below(ws, delete_at)

        ws.delete_rows(delete_at, deleted)

        # Remove any phantom ranges left at the deletion point
        stale = [rng for rng in ws.merged_cells.ranges
                 if delete_at <= rng.min_row < delete_at + deleted]
        for rng in stale:
            ws.merged_cells.ranges.remove(rng)

        _reapply_merges(ws, saved, -deleted)

    shift = offset - deleted

    # ── Project (global, one per invoice) ───────────────────────────────────
    is_led = (project == 'LED灯具')

    # ── Fill item rows (rows 17 onward) ──────────────────────────────────────
    for i in range(n):
        r = ITEMS_START + i
        p = str(i)
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = form.get(f'item_desc_{p}',   '')
        ws.cell(r, 3).value = form.get(f'item_model_{p}',  '')
        if is_led:
            dia_shape  = form.get(f'item_dia_shape_{p}',  '')
            lumin      = form.get(f'item_lumin_{p}',      '').rstrip('lm').strip()
            voltage    = form.get(f'item_voltage_{p}',    '').rstrip('Vv').strip()
            power      = form.get(f'item_power_{p}',      '').rstrip('Ww').strip()
            color_temp = form.get(f'item_color_temp_{p}', '').rstrip('Kk').strip()
            material   = form.get(f'item_material_{p}',   '')
            ws.cell(r, 4).value = material if material else ''
            ws.cell(r, 5).value = dia_shape
            ws.cell(r, 6).value = f'{lumin}lm' if lumin else ''
            ws.cell(r, 11).value = f'{voltage}V' if voltage else ''
            ws.cell(r, 12).value = f'{power}W' if power else ''
            ws.cell(r, 13).value = f'{color_temp}K' if color_temp else ''
        else:
            ws.cell(r, 5).value = form.get(f'item_color_{p}',  '')
            ws.cell(r, 6).value = form.get(f'item_size_{p}',   '')
        ws.cell(r, 7).value = form.get(f'item_unit_{p}',   'Set')
        try:    qty   = float(form.get(f'item_qty_{p}')   or 0)
        except: qty   = 0
        try:    price = float(form.get(f'item_price_{p}') or 0)
        except: price = 0
        ws.cell(r, 8).value  = qty
        ws.cell(r, 9).value  = price
        ws.cell(r, 10).value = f'=H{r}*I{r}'
        tax_raw  = form.get(f'item_tax_{p}', 'excl')
        if is_ar:
            tax_note = 'Incl. 16% GST شامل الضريبة' if tax_raw == 'incl' else 'Excl. Tax غير شامل الضريبة'
        else:
            tax_note = 'Incl. 16% GST 含16%消费税' if tax_raw == 'incl' else 'Excl. Tax 不含税'
        remarks  = (form.get(f'item_remarks_{p}') or '/').strip()
        ws.cell(r, 14).value = tax_note if remarks == '/' else f'{remarks} [{tax_note}]'

    # ── Dynamic column headers for LED project (row 16) ──────────────────
    if is_led:
        ws.cell(16, 4).value = 'Material'
        ws.cell(16, 5).value = 'Dia / Shape'
        ws.cell(16, 6).value = 'Lumin (lm)'
        ws.cell(16, 11).value = 'Voltage'
        ws.cell(16, 12).value = 'Power'
        ws.cell(16, 13).value = 'Color Temp'

    # ── Column widths & orientation ─────────────────────────────────────
    if is_led:
        # Landscape for LED: wider columns, all 14 visible
        ws.page_setup.orientation = 'landscape'
        for col in ('D', 'K', 'L', 'M'):
            ws.column_dimensions[col].hidden = False
        widths = {'A':5,'B':42,'C':14,'D':13,'E':13,'F':13,'G':10,'H':9,
                  'I':15,'J':15,'K':13,'L':12,'M':13,'N':15}
    else:
        for col in ('D', 'K', 'L', 'M'):
            ws.column_dimensions[col].hidden = True
        widths = {'A':4,'B':32,'C':11,'E':11,'F':11,'G':8,'H':7,
                  'I':12,'J':12,'N':12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Summary rows (Subtotal=22, Freight=23, Grand=24, base) ───────────────
    last_item = ITEMS_START + n - 1
    sr = 22 + shift
    fr = 23 + shift
    gr = 24 + shift

    col_j = get_column_letter(10)

    ws.cell(sr, 10).value = f'=SUM({col_j}{ITEMS_START}:{col_j}{last_item})'
    try:    freight = float(form.get('freight') or 0)
    except: freight = 0
    ws.cell(fr, 10).value = freight
    ws.cell(gr, 10).value = f'={col_j}{sr}+{col_j}{fr}'
    ws.cell(gr, 14).value = currency

    # ── Sync validity into Remarks item 4 (base row 29); skip for LED ────
    if not is_led:
        remarks_row4 = 29 + shift
        cell_r4 = ws.cell(remarks_row4, 1)
        if cell_r4.value and '[  ]' in str(cell_r4.value):
            replacement = validity if validity else '___'
            cell_r4.value = str(cell_r4.value).replace('[  ]', replacement)

    # ── LED-specific remarks (replace all standard remarks rows) ───────────
    if is_led:
        base = 26 + shift
        # Unmerge remarks rows before writing
        for rr in range(base, base + 5):
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row <= rr <= merged_range.max_row:
                    ws.unmerge_cells(str(merged_range))
        months = (form.get('led_delivery_months') or '').strip()
        deposit_pct = (form.get('led_deposit_pct') or '').strip()
        remaining_pct = (form.get('led_remaining_pct') or '').strip()
        ws.cell(base,     1).value = '1. The above prices are ex-factory prices.'
        ws.cell(base + 1, 1).value = f'2. Delivery time: {months} months after receipt of {deposit_pct}% deposit.'
        ws.cell(base + 2, 1).value = f'3. The remaining: {remaining_pct}% Will be paid upon delivery.'
        ws.cell(base + 3, 1).value = '4. Warranty: Bulb-1year.'
        ws.cell(base + 4, 1).value = None  # Clear 5th remark row (LED only has 4)
        # Merge each LED remark row across all columns
        for rr in range(base, base + 4):
            ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=14)

    # ── Bank / transfer details (rows 32-35) ────────────────────────────────
    bank_project = form.get('selected_bank_project', '').strip()
    if bank_project:
        all_banks = load_bank_info()
        bank = next((b for b in all_banks if b.get('project') == bank_project), None)
        if bank:
            br = 32 + shift
            # Unmerge bank rows first (delete_rows may have corrupted merged ranges)
            for rr in range(br, br + 4):
                for merged_range in list(ws.merged_cells.ranges):
                    if merged_range.min_row <= rr <= merged_range.max_row:
                        ws.unmerge_cells(str(merged_range))
            # Now write values safely
            ws.cell(br,     3).value = bank.get('beneficiary', '')
            ws.cell(br,    10).value = bank.get('bank', '')
            ws.cell(br + 1, 3).value = bank.get('swift', '')
            ws.cell(br + 1,10).value = bank.get('account', '')
            ws.cell(br + 2, 3).value = bank.get('iban', '')
            ws.cell(br + 3, 3).value = bank.get('bank_address', '')

    # ── CI origin statement (row 37) ─────────────────────────────────────────
    if mode == 'ci':
        origin_row = 37 + shift
        ws.merge_cells(start_row=origin_row, start_column=1, end_row=origin_row, end_column=14)
        ws.cell(origin_row, 1).value = (
            'STATEMENT : we hereby certify that this invoice is in all respect correct and true '
            'both with regard to the price and description of the goods referred herein, '
            'and that the country of origin is Jordan'
        )
        ws.cell(origin_row, 1).font = openpyxl.styles.Font(bold=True, italic=True, size=11)
        ws.cell(origin_row, 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # ── Company logo (A1:D2 area) ────────────────────────────────────────────
    selected_logo = form.get('selected_logo_name', '').strip()
    if selected_logo:
        logo_path = os.path.join(LOGO_DIR, selected_logo)
        if os.path.exists(logo_path):
            ws.cell(1, 1).value = None
            with open(logo_path, 'rb') as lf:
                insert_image(ws, lf, 'A1', w=150, h=70)

    # ── Update print area to reflect new last row ────────────────────────────
    last_row = (37 if mode == 'ci' else 36) + shift
    ws.print_area = f'A1:N{last_row}'

    # ── Output ────────────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    inv_no   = (form.get('invoice_no') or 'PI').strip()
    customer = (form.get('customer_name') or '').strip()
    suffix   = '_AR' if is_ar else ''
    prefix   = 'CI' if mode == 'ci' else 'PI'
    filename = f"{prefix}_{inv_no}_{customer}{suffix}.xlsx".replace(' ', '_').replace('/', '-')

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Ceramic Tile PI ─────────────────────────────────────────────────────────────

CERAMIC_ITEMS_START = 18
CERAMIC_ITEMS_END   = 33
CERAMIC_ITEMS_SLOTS = 16


def fill_ceramic_pi(form, files, lang='zh', mode='pi'):
    with open(CERAMIC_PI_TEMPLATE_PATH, 'rb') as f:
        buf = io.BytesIO(f.read())
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    is_ar = (lang == 'ar')
    n = max(1, int(form.get('item_count') or 1))

    # Supplier info (rows 4-8, col C)
    ws.cell(4, 3).value = form.get('company_name', '')
    ws.cell(5, 3).value = form.get('address', '')
    ws.cell(6, 3).value = form.get('city_country', '')
    tel = (form.get('tel') or '').strip()
    ws.cell(7, 3).value = tel if tel else None
    ws.cell(8, 3).value = form.get('email', '')

    # Invoice info (rows 4-8, col J)
    ws.cell(4, 10).value = form.get('invoice_no', '')
    date_str = format_invoice_date(form.get('date', ''))
    ws.cell(5, 10).value = date_str
    currency = form.get('currency') or 'JOD'
    ws.cell(6, 10).value = currency
    ws.cell(7, 10).value = form.get('incoterms', '')
    ws.cell(8, 10).value = form.get('payment_terms', '')

    # Customer info (rows 10-14, col C)
    ws.cell(10, 3).value = form.get('customer_name', '')
    ws.cell(11, 3).value = form.get('customer_address', '')
    ws.cell(12, 3).value = form.get('customer_email', '')
    ws.cell(13, 3).value = form.get('country', '')
    cust_tel = (form.get('customer_tel') or '').strip()
    ws.cell(14, 3).value = cust_tel if cust_tel else None

    # Quotation details (rows 10-14, col J)
    prefix = (form.get('contact_prefix') or '').strip()
    person = (form.get('contact_person')  or '').strip()
    ws.cell(10, 10).value = f"{prefix} {person}".strip()
    ws.cell(11, 10).value = form.get('inquiry_ref', '')
    ws.cell(12, 10).value = form.get('lead_time', '')
    validity = (form.get('quotation_validity') or '').strip()
    ws.cell(13, 10).value = f'{validity} days' if validity else ''
    port = (form.get('port_of_loading') or '').strip()
    dest = (form.get('destination') or '').strip()
    if port and dest:    shipping = f'{port} → {dest}'
    elif port:          shipping = port
    else:               shipping = dest
    ws.cell(14, 10).value = shipping

    # Arabic headers
    if is_ar:
        ws.cell(3,  1).value = 'FROM / SUPPLIER  المورّد'
        ws.cell(3,  8).value = 'INVOICE INFO  معلومات الفاتورة'
        ws.cell(9,  1).value = 'TO / CUSTOMER  العميل'
        ws.cell(9,  8).value = 'QUOTATION DETAILS  تفاصيل العرض'
        ws.cell(16, 1).value = 'ITEM LIST  قائمة المنتجات — Ceramic Tiles / بلاط السيراميك'
        ws.cell(37, 1).value = 'REMARKS  ملاحظات'
        ws.cell(43, 1).value = 'TRANSFER DETAILS  معلومات التحويل'
        ws.cell(48, 1).value = 'Company Stamp  ختم الشركة:'

    # CI title override
    if mode == 'ci':
        title_cell = ws.cell(1, 8)
        if title_cell.value and 'PROFORMA' in str(title_cell.value):
            title_cell.value = str(title_cell.value).replace('PROFORMA INVOICE', 'COMMERCIAL INVOICE')

    # Insert / delete item rows (16 slots, rows 18-33)
    offset, deleted, shift = shift_item_rows(ws, CERAMIC_ITEMS_START, CERAMIC_ITEMS_END, n)

    # Fill item rows (tile-specific 14 columns)
    for i in range(n):
        r = CERAMIC_ITEMS_START + i
        p = str(i)
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = form.get(f'item_model_{p}',  '')
        ws.cell(r, 3).value = form.get(f'item_tile_type_{p}', '')
        ws.cell(r, 4).value = form.get(f'item_size_{p}',   '')
        ws.cell(r, 5).value = form.get(f'item_thickness_{p}', '')
        ws.cell(r, 6).value = form.get(f'item_brand_{p}',  '')
        ws.cell(r, 7).value = form.get(f'item_m2_per_ctn_{p}', '')
        ws.cell(r, 8).value = form.get(f'item_gw_per_ctn_{p}', '')
        ws.cell(r, 9).value = form.get(f'item_nw_per_ctn_{p}', '')
        ws.cell(r, 10).value = form.get(f'item_unit_{p}',  'm²')
        try:    qty   = float(form.get(f'item_qty_{p}')   or 0)
        except: qty   = 0
        try:    price = float(form.get(f'item_price_{p}') or 0)
        except: price = 0
        ws.cell(r, 11).value = qty
        ws.cell(r, 12).value = price
        ws.cell(r, 13).value = f'=K{r}*L{r}'
        tax_raw  = form.get(f'item_tax_{p}', 'excl')
        if is_ar:
            tax_note = 'Incl. 16% GST شامل الضريبة' if tax_raw == 'incl' else 'Excl. Tax غير شامل الضريبة'
        else:
            tax_note = 'Incl. 16% GST 含16%消费税' if tax_raw == 'incl' else 'Excl. Tax 不含税'
        remarks  = (form.get(f'item_remarks_{p}') or '/').strip()
        ws.cell(r, 14).value = tax_note if remarks == '/' else f'{remarks} [{tax_note}]'

    # Summary rows (34-36)
    last_item = CERAMIC_ITEMS_START + n - 1
    sr = 34 + shift
    fr = 35 + shift
    gr = 36 + shift

    col_m = get_column_letter(13)
    ws.cell(sr, 13).value = f'=SUM({col_m}{CERAMIC_ITEMS_START}:{col_m}{last_item})'
    try:    freight = float(form.get('freight') or 0)
    except: freight = 0
    ws.cell(fr, 13).value = freight
    ws.cell(gr, 13).value = f'={col_m}{sr}+{col_m}{fr}'
    ws.cell(gr, 14).value = currency

    # Remarks: sync validity into row 4
    remarks_row4 = 41 + shift
    cell_r4 = ws.cell(remarks_row4, 1)
    if cell_r4.value and '[  ]' in str(cell_r4.value):
        replacement = validity if validity else '___'
        cell_r4.value = str(cell_r4.value).replace('[  ]', replacement)

    # Bank info (rows 44-47)
    bank_project = form.get('selected_bank_project', '').strip()
    if bank_project:
        all_banks = load_bank_info()
        bank = next((b for b in all_banks if b.get('project') == bank_project), None)
        if bank:
            fill_bank_from_json(ws, 44 + shift, bank)

    # CI origin statement (row 49)
    if mode == 'ci':
        origin_row = 49 + shift
        ws.merge_cells(start_row=origin_row, start_column=1, end_row=origin_row, end_column=14)
        ws.cell(origin_row, 1).value = (
            'STATEMENT : we hereby certify that this invoice is in all respect correct and true '
            'both with regard to the price and description of the goods referred herein, '
            'and that the country of origin is Jordan'
        )
        ws.cell(origin_row, 1).font = openpyxl.styles.Font(bold=True, italic=True, size=11)
        ws.cell(origin_row, 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # Logo
    selected_logo = form.get('selected_logo_name', '').strip()
    if selected_logo:
        logo_path = os.path.join(LOGO_DIR, selected_logo)
        if os.path.exists(logo_path):
            ws.cell(1, 1).value = None
            with open(logo_path, 'rb') as lf:
                excel_insert_image(ws, lf, 'A1', w=150, h=70)

    # Print area
    last_row = (49 if mode == 'ci' else 48) + shift
    ws.print_area = f'A1:N{last_row}'

    # Output
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    inv_no   = (form.get('invoice_no') or 'PI').strip()
    customer = (form.get('customer_name') or '').strip()
    suffix   = '_AR' if is_ar else ''
    prefix   = 'CI' if mode == 'ci' else 'PI'
    filename = f"{prefix}_{inv_no}_{customer}{suffix}.xlsx".replace(' ', '_').replace('/', '-')

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Preview context builder (shared by PI and CI) ──────────────────────────────

def build_preview_context(form, files, mode='pi'):
    n = max(1, int(form.get('item_count') or 1))

    project = form.get('project', '').strip()
    is_led = (project == 'LED灯具')
    is_tile = (project == '瓷砖')

    items = []
    for i in range(n):
        p = str(i)
        try:    qty   = float(form.get(f'item_qty_{p}')   or 0)
        except: qty   = 0
        try:    price = float(form.get(f'item_price_{p}') or 0)
        except: price = 0
        tax_raw  = form.get(f'item_tax_{p}', 'excl')
        tax_note = 'Incl. 16% GST 含16%消费税' if tax_raw == 'incl' else 'Excl. Tax 不含税'
        remarks  = (form.get(f'item_remarks_{p}') or '/').strip()
        remark_combined = tax_note if remarks == '/' else f'{remarks} [{tax_note}]'
        item = {
            'no':       i + 1,
            'desc':     form.get(f'item_desc_{p}',    ''),
            'model':    form.get(f'item_model_{p}',   ''),
            'unit':     form.get(f'item_unit_{p}',    'Set'),
            'qty':      qty,
            'price':    price,
            'amount':   qty * price,
            'remarks':  remark_combined,
            'is_led':   is_led,
        }
        if is_led:
            item['dia_shape']  = form.get(f'item_dia_shape_{p}',  '')
            item['lumin']      = form.get(f'item_lumin_{p}',      '').rstrip('lm').strip()
            item['voltage']    = form.get(f'item_voltage_{p}',    '').rstrip('Vv').strip()
            item['power']      = form.get(f'item_power_{p}',      '').rstrip('Ww').strip()
            item['color_temp'] = form.get(f'item_color_temp_{p}', '').rstrip('Kk').strip()
            item['material']   = form.get(f'item_material_{p}',   '')
        elif is_tile:
            item['tile_type']  = form.get(f'item_tile_type_{p}',  '')
            item['size']       = form.get(f'item_size_{p}',       '')
            item['thickness']  = form.get(f'item_thickness_{p}',  '')
            item['brand']      = form.get(f'item_brand_{p}',      '')
            item['m2_per_ctn'] = form.get(f'item_m2_per_ctn_{p}', '')
            item['gw_per_ctn'] = form.get(f'item_gw_per_ctn_{p}', '')
            item['nw_per_ctn'] = form.get(f'item_nw_per_ctn_{p}', '')
        else:
            item['color'] = form.get(f'item_color_{p}', '')
            item['size']  = form.get(f'item_size_{p}',  '')
        items.append(item)

    try:    freight = float(form.get('freight') or 0)
    except: freight = 0
    subtotal = sum(it['amount'] for it in items)
    grand    = subtotal + freight

    currency = form.get('currency') or 'JOD'
    date_str = format_invoice_date(form.get('date', ''))

    prefix = (form.get('contact_prefix') or '').strip()
    person = (form.get('contact_person')  or '').strip()
    contact = f"{prefix} {person}".strip()

    port = (form.get('port_of_loading') or '').strip()
    dest = (form.get('destination') or '').strip()
    if port and dest:    shipping = f'{port} → {dest}'
    elif port:          shipping = port
    else:               shipping = dest

    # Logo as base64
    logo_b64 = ''
    logo_mime = 'image/png'
    selected_logo = form.get('selected_logo_name', '').strip()
    if selected_logo:
        logo_path = os.path.join(LOGO_DIR, selected_logo)
        if os.path.exists(logo_path):
            ext = os.path.splitext(selected_logo)[1].lower()
            mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg', '.gif': 'image/gif', '.webp': 'image/webp'}
            logo_mime = mime_map.get(ext, 'image/png')
            with open(logo_path, 'rb') as lf:
                logo_b64 = base64.b64encode(lf.read()).decode()

    lang = form.get('lang', 'zh')
    is_ar = (lang == 'ar')

    # Rebuild items with language-aware tax notes
    for it in items:
        p = str(it['no'] - 1)
        tax_raw = form.get(f'item_tax_{p}', 'excl')
        if is_ar:
            tax_note = 'Incl. 16% GST شامل الضريبة' if tax_raw == 'incl' else 'Excl. Tax غير شامل الضريبة'
        else:
            tax_note = 'Incl. 16% GST 含16%消费税' if tax_raw == 'incl' else 'Excl. Tax 不含税'
        remarks = (form.get(f'item_remarks_{p}') or '/').strip()
        it['remarks'] = tax_note if remarks == '/' else f'{remarks} [{tax_note}]'

    invoice_type = 'COMMERCIAL INVOICE' if mode == 'ci' else None
    origin_statement = (
        'STATEMENT : we hereby certify that this invoice is in all respect correct and true '
        'both with regard to the price and description of the goods referred herein, '
        'and that the country of origin is Jordan'
    ) if mode == 'ci' else None

    ctx = {
        'company_name':     form.get('company_name', ''),
        'address':          form.get('address', ''),
        'city_country':     form.get('city_country', ''),
        'tel':              form.get('tel', ''),
        'email':            form.get('email', ''),
        'invoice_no':       form.get('invoice_no', ''),
        'date_str':         date_str,
        'customer_name':    form.get('customer_name', ''),
        'customer_address': form.get('customer_address', ''),
        'customer_email':   form.get('customer_email', ''),
        'country':          form.get('country', ''),
        'customer_tel':     form.get('customer_tel', ''),
        'contact':          contact,
        'inquiry_ref':      form.get('inquiry_ref', ''),
        'currency':         currency,
        'incoterms':        form.get('incoterms', ''),
        'payment_terms':    form.get('payment_terms', ''),
        'lead_time':        form.get('lead_time', ''),
        'validity':         form.get('quotation_validity', ''),
        'shipping':         shipping,
        'items':            items,
        'subtotal':         subtotal,
        'freight':          freight,
        'grand':            grand,
        'logo_b64':         logo_b64,
        'logo_mime':        logo_mime,
        'bank':             None,
        'is_led':           is_led,
        'is_tile':          is_tile,
        'project':          project,
        'invoice_type':     invoice_type,
        'origin_statement': origin_statement,
        'led_delivery_months': form.get('led_delivery_months', '').strip(),
        'led_deposit_pct':     form.get('led_deposit_pct',     '').strip(),
        'led_remaining_pct':   form.get('led_remaining_pct',   '').strip(),
    }
    bank_project = form.get('selected_bank_project', '').strip()
    if bank_project:
        all_banks = load_bank_info()
        bank = next((b for b in all_banks if b.get('project') == bank_project), None)
        if bank:
            ctx['bank'] = bank
    return ctx


# ── PL Preview context builder ─────────────────────────────────────────────────

def build_pl_preview_context(form, files):
    n = max(1, int(form.get('item_count') or 1))
    project = form.get('project', '').strip()
    is_tile = (project == '瓷砖')

    items = []
    total_ctns = 0
    total_gw = 0
    total_nw = 0
    total_cbm = 0
    for i in range(n):
        p = str(i)
        try:    qty = float(form.get(f'item_qty_{p}') or 0)
        except: qty = 0

        if is_tile:
            try:    gw = float(form.get(f'item_gw_per_ctn_{p}') or 0)
            except: gw = 0
            try:    nw = float(form.get(f'item_nw_per_ctn_{p}') or 0)
            except: nw = 0
            try:    m2 = float(form.get(f'item_m2_per_ctn_{p}') or 1)
            except: m2 = 1
            ctns = qty / m2 if m2 > 0 else 0
            total_ctns += ctns
            total_gw += ctns * gw
            total_nw += ctns * nw
            item = {
                'no':         i + 1,
                'model':      form.get(f'item_model_{p}',       ''),
                'tile_type':  form.get(f'item_tile_type_{p}',   ''),
                'size':       form.get(f'item_size_{p}',        ''),
                'thickness':  form.get(f'item_thickness_{p}',   ''),
                'brand':      form.get(f'item_brand_{p}',       ''),
                'm2_per_ctn': form.get(f'item_m2_per_ctn_{p}',  ''),
                'gw_per_ctn': form.get(f'item_gw_per_ctn_{p}',  ''),
                'nw_per_ctn': form.get(f'item_nw_per_ctn_{p}',  ''),
                'unit':       form.get(f'item_unit_{p}',        'm²'),
                'qty':        qty,
                'ctn_no':     (form.get(f'item_ctn_no_{p}') or '').strip(),
                'remarks':    (form.get(f'item_remarks_{p}') or '/').strip(),
            }
        else:
            try:    pcs_per_ctn = float(form.get(f'item_pcs_per_ctn_{p}') or 1)
            except: pcs_per_ctn = 1
            try:    gw_per_ctn  = float(form.get(f'item_gw_per_ctn_{p}') or 0)
            except: gw_per_ctn  = 0
            try:    nw_per_ctn  = float(form.get(f'item_nw_per_ctn_{p}') or 0)
            except: nw_per_ctn  = 0
            ctns = qty / pcs_per_ctn if pcs_per_ctn > 0 else 0
            gw = ctns * gw_per_ctn
            nw = ctns * nw_per_ctn
            carton_size = (form.get(f'item_carton_size_{p}') or '').strip()
            cbm = 0
            dims = re.match(r'(\d+\.?\d*)\s*[×xX]\s*(\d+\.?\d*)\s*[×xX]\s*(\d+\.?\d*)', carton_size)
            if dims:
                vol_cm3 = float(dims[1]) * float(dims[2]) * float(dims[3])
                cbm = ctns * vol_cm3 / 1000000
            total_ctns += ctns
            total_gw += gw
            total_nw += nw
            total_cbm += cbm
            item = {
                'no':           i + 1,
                'desc':         form.get(f'item_desc_{p}',        ''),
                'model':        form.get(f'item_model_{p}',       ''),
                'qty':          qty,
                'unit':         form.get(f'item_unit_{p}',        'Set'),
                'ctns':         ctns,
                'pcs_per_ctn':  pcs_per_ctn,
                'gw':           gw,
                'nw':           nw,
                'carton_size':  carton_size,
                'cbm':          cbm,
                'ctn_no':       (form.get(f'item_ctn_no_{p}') or '').strip(),
                'remarks':      (form.get(f'item_remarks_{p}') or '/').strip(),
            }
        items.append(item)

    transport_mode = (form.get('transport_mode') or 'sea').strip()
    is_sea = (transport_mode == 'sea')
    transport_label = 'Sea 海运' if is_sea else 'Land 陆运'
    port_label = 'Port of Loading' if is_sea else 'Loading Place'
    port_label_ar = 'Port of Loading 装运港' if is_sea else 'Loading Place 装运地'
    ctn_header = 'CTN No.<br>柜号' if is_sea else 'Truck Plate<br>车牌号'
    ctn_header_ar = 'CTN No.<br>رقم الحاوية' if is_sea else 'Truck Plate<br>رقم الشاحنة'

    date_str = format_invoice_date(form.get('date', ''))

    prefix = (form.get('contact_prefix') or '').strip()
    person = (form.get('contact_person')  or '').strip()
    contact = f"{prefix} {person}".strip()

    port = (form.get('port_of_loading') or '').strip()
    dest = (form.get('destination') or '').strip()

    logo_b64 = ''
    logo_mime = 'image/png'
    selected_logo = form.get('selected_logo_name', '').strip()
    if selected_logo:
        logo_path = os.path.join(LOGO_DIR, selected_logo)
        if os.path.exists(logo_path):
            ext = os.path.splitext(selected_logo)[1].lower()
            mime_map = {'.png': 'image/png', '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg', '.gif': 'image/gif', '.webp': 'image/webp'}
            logo_mime = mime_map.get(ext, 'image/png')
            with open(logo_path, 'rb') as lf:
                logo_b64 = base64.b64encode(lf.read()).decode()

    ctx = {
        'company_name':     form.get('company_name', ''),
        'address':          form.get('address', ''),
        'city_country':     form.get('city_country', ''),
        'tel':              form.get('tel', ''),
        'email':            form.get('email', ''),
        'invoice_no':       form.get('invoice_no', ''),
        'date_str':         date_str,
        'customer_name':    form.get('customer_name', ''),
        'customer_address': form.get('customer_address', ''),
        'customer_email':   form.get('customer_email', ''),
        'country':          form.get('country', ''),
        'customer_tel':     form.get('customer_tel', ''),
        'contact':          contact,
        'inquiry_ref':      form.get('inquiry_ref', ''),
        'invoice_ref':      form.get('invoice_ref', ''),
        'port':             port,
        'destination':      dest,
        'transport_mode':   transport_mode,
        'transport_label':  transport_label,
        'port_label':       port_label,
        'port_label_ar':    port_label_ar,
        'ctn_header':       ctn_header,
        'ctn_header_ar':    ctn_header_ar,
        'is_tile':          is_tile,
        'items':            items,
        'total_ctns':       total_ctns,
        'total_gw':         total_gw,
        'total_nw':         total_nw,
        'total_cbm':        total_cbm,
        'logo_b64':         logo_b64,
        'logo_mime':        logo_mime,
    }
    return ctx


# ── PL Excel generation ────────────────────────────────────────────────────────

def fill_pl_template(form, files, lang='zh'):
    is_ar = (lang == 'ar')
    n = max(1, int(form.get('item_count') or 1))
    project = form.get('project', '').strip()
    is_tile = (project == '瓷砖')
    transport_mode = (form.get('transport_mode') or 'sea').strip()
    is_sea = (transport_mode == 'sea')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PACKING LIST'

    # Page setup: A4 landscape
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_margins.left = 0.24
    ws.page_margins.right = 0.24
    ws.page_margins.top = 0.24
    ws.page_margins.bottom = 0.24

    # Column widths (13 columns; M hidden as spare)
    if is_tile:
        widths = {'A':2.8, 'B':12, 'C':6, 'D':7, 'E':4.5, 'F':7, 'G':6,
                  'H':6.5, 'I':6.5, 'J':5, 'K':6.5, 'L':10, 'N':20.2}
    else:
        widths = {'A':2.8, 'B':15, 'C':8, 'D':5, 'E':5, 'F':5, 'G':5,
                  'H':5.5, 'I':5.5, 'J':9, 'K':6, 'L':10, 'N':17.2}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions['M'].hidden = True

    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin'),
    )
    thick_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='medium'),
        right=openpyxl.styles.Side(style='medium'),
        top=openpyxl.styles.Side(style='medium'),
        bottom=openpyxl.styles.Side(style='medium'),
    )

    row = 1

    # Row 1-2: Logo + Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row+1, end_column=14)
    title_cell = ws.cell(row, 5)
    title_cell.value = 'PACKING LIST\n装箱单' if not is_ar else 'PACKING LIST\nقائمة التعبئة'
    title_cell.font = openpyxl.styles.Font(bold=True, size=26)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    for c in range(1, 15):
        for rr in range(row, row+2):
            ws.cell(rr, c).border = thin_border
            ws.cell(rr, c).font = openpyxl.styles.Font(size=10)
    ws.row_dimensions[row].height = 14
    ws.row_dimensions[row+1].height = 14
    row = 3

    # Row 3: Section bars
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=14)
    c1 = ws.cell(row, 1)
    c2 = ws.cell(row, 8)
    c1.value = 'FROM / SUPPLIER  供货商' if not is_ar else 'FROM / SUPPLIER  المورّد'
    c2.value = 'PACKING LIST INFO  装箱单信息' if not is_ar else 'PACKING LIST INFO  معلومات قائمة التعبئة'
    for c in [c1, c2]:
        c.font = openpyxl.styles.Font(bold=True, size=12)
        c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        c.border = thick_border
    for c in range(1, 15):
        ws.cell(row, c).border = thick_border
    ws.row_dimensions[row].height = 20
    row = 4

    # Helper: write a label+value row
    def write_info_row(r, ll, lv, rl, rv):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=10)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=14)
        ws.cell(r, 1).value = ll
        ws.cell(r, 1).font = openpyxl.styles.Font(bold=True, size=10)
        ws.cell(r, 3).value = lv
        ws.cell(r, 3).font = openpyxl.styles.Font(bold=True, size=11)
        ws.cell(r, 8).value = rl
        ws.cell(r, 8).font = openpyxl.styles.Font(bold=True, size=10)
        ws.cell(r, 11).value = rv
        ws.cell(r, 11).font = openpyxl.styles.Font(bold=True, size=11)
        for c in range(1, 15):
            ws.cell(r, c).border = thin_border
        ws.row_dimensions[r].height = 19

    # Supplier + PL Info rows (rows 4-8)
    date_str = format_invoice_date(form.get('date', ''))
    transport_label = 'Sea 海运' if is_sea else 'Land 陆运'
    info_rows = [
        ('Company Name:', form.get('company_name', ''), 'PL No.:', form.get('invoice_no', '')),
        ('Address:', form.get('address', ''), 'Date:', date_str),
        ('City / Country:', form.get('city_country', ''), 'Invoice Ref.:', form.get('invoice_ref', '')),
        ('Tel:', form.get('tel', ''), 'Transport Mode:', transport_label),
        ('Email:', form.get('email', ''), '', ''),
    ]
    for ll, lv, rl, rv in info_rows:
        write_info_row(row, ll, lv, rl, rv)
        row += 1

    # Row 9: Customer + Shipping bars
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=14)
    c1 = ws.cell(row, 1)
    c2 = ws.cell(row, 8)
    c1.value = 'TO / CUSTOMER  客户' if not is_ar else 'TO / CUSTOMER  العميل'
    c2.value = 'SHIPPING DETAILS  运输详情' if not is_ar else 'SHIPPING DETAILS  تفاصيل الشحن'
    for c in [c1, c2]:
        c.font = openpyxl.styles.Font(bold=True, size=12)
        c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        c.border = thick_border
    for c in range(1, 15):
        ws.cell(row, c).border = thick_border
    ws.row_dimensions[row].height = 20
    row = 10

    # Customer + Shipping Details (rows 10-14)
    prefix = (form.get('contact_prefix') or '').strip()
    person = (form.get('contact_person')  or '').strip()
    contact = f"{prefix} {person}".strip()
    port = (form.get('port_of_loading') or '').strip()
    dest = (form.get('destination') or '').strip()
    port_lbl = 'Port of Loading' if is_sea else 'Loading Place'
    cust_rows = [
        ('Customer Name:', form.get('customer_name', ''), 'Contact Person:', contact),
        ('Customer Address:', form.get('customer_address', ''), 'Project / Inquiry Ref.:', form.get('inquiry_ref', '')),
        ('Country:', form.get('country', ''), f'{port_lbl}:', port),
        ('Customer Tel:', form.get('customer_tel', ''), 'Destination:', dest),
        ('Customer Email:', form.get('customer_email', ''), 'Country of Origin:', 'Jordan'),
    ]
    for ll, lv, rl, rv in cust_rows:
        write_info_row(row, ll, lv, rl, rv)
        row += 1

    # Spacer
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.row_dimensions[row].height = 7
    row = 16

    # ITEM LIST bar
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    c = ws.cell(row, 1)
    c.value = 'ITEM LIST  商品明细 — PACKING LIST / 装箱单' if not is_ar else 'ITEM LIST  قائمة المنتجات — PACKING LIST / قائمة التعبئة'
    c.font = openpyxl.styles.Font(bold=True, size=12)
    c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    c.border = thick_border
    for cc in range(1, 15):
        ws.cell(row, cc).border = thick_border
    ws.row_dimensions[row].height = 20
    row = 17

    # Column headers
    ctn_header = 'CTN No.\n柜号' if is_sea else 'Truck Plate\n车牌号'
    if is_tile:
        headers = [
            'Item\nNo.', 'ITEM NO.', 'TYPE', 'PALLET SIZE', 'THK\n(mm)', 'BRAND',
            'M²/CTN', 'GW/CTN\n(kg)', 'NW/CTN\n(kg)', 'Unit', 'Qty\n(m²)',
            ctn_header, 'Remarks',
        ]
    else:
        headers = [
            'Item\nNo.', 'Product Description', 'Product No.', 'Qty', 'Unit',
            'CTNS', 'Pcs/CTN', 'GW\n(kg)', 'NW\n(kg)', 'Carton Size\n(cm)',
            'CBM\n(m³)', ctn_header, 'Remarks',
        ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i)
        c.value = h
        c.font = openpyxl.styles.Font(bold=True, size=9)
        c.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thick_border
    ws.row_dimensions[row].height = 24
    row = 18

    # Item data
    total_ctns = 0
    total_gw = 0
    total_nw = 0
    total_cbm = 0
    for i in range(n):
        p = str(i)
        r = row + i
        try:    qty = float(form.get(f'item_qty_{p}') or 0)
        except: qty = 0

        if is_tile:
            try:    m2 = float(form.get(f'item_m2_per_ctn_{p}') or 1)
            except: m2 = 1
            try:    gw = float(form.get(f'item_gw_per_ctn_{p}') or 0)
            except: gw = 0
            try:    nw = float(form.get(f'item_nw_per_ctn_{p}') or 0)
            except: nw = 0
            ctns = qty / m2 if m2 > 0 else 0
            total_ctns += ctns
            total_gw += ctns * gw
            total_nw += ctns * nw
            vals = [
                i + 1,
                form.get(f'item_model_{p}', ''),
                form.get(f'item_tile_type_{p}', ''),
                form.get(f'item_size_{p}', ''),
                form.get(f'item_thickness_{p}', ''),
                form.get(f'item_brand_{p}', ''),
                form.get(f'item_m2_per_ctn_{p}', ''),
                form.get(f'item_gw_per_ctn_{p}', ''),
                form.get(f'item_nw_per_ctn_{p}', ''),
                form.get(f'item_unit_{p}', 'm²'),
                qty,
                (form.get(f'item_ctn_no_{p}') or '').strip(),
                (form.get(f'item_remarks_{p}') or '/').strip(),
            ]
            num_cols = (7, 8, 9, 11)
        else:
            try:    pcs_per_ctn = float(form.get(f'item_pcs_per_ctn_{p}') or 1)
            except: pcs_per_ctn = 1
            try:    gw_per_ctn  = float(form.get(f'item_gw_per_ctn_{p}') or 0)
            except: gw_per_ctn  = 0
            try:    nw_per_ctn  = float(form.get(f'item_nw_per_ctn_{p}') or 0)
            except: nw_per_ctn  = 0
            ctns = qty / pcs_per_ctn if pcs_per_ctn > 0 else 0
            gw = ctns * gw_per_ctn
            nw = ctns * nw_per_ctn
            carton_size = (form.get(f'item_carton_size_{p}') or '').strip()
            cbm = 0
            dims = re.match(r'(\d+\.?\d*)\s*[×xX]\s*(\d+\.?\d*)\s*[×xX]\s*(\d+\.?\d*)', carton_size)
            if dims:
                vol_cm3 = float(dims[1]) * float(dims[2]) * float(dims[3])
                cbm = ctns * vol_cm3 / 1000000
            total_ctns += ctns
            total_gw += gw
            total_nw += nw
            total_cbm += cbm
            vals = [
                i + 1,
                form.get(f'item_desc_{p}', ''),
                form.get(f'item_model_{p}', ''),
                qty,
                form.get(f'item_unit_{p}', 'Set'),
                ctns,
                pcs_per_ctn,
                gw,
                nw,
                carton_size,
                cbm,
                (form.get(f'item_ctn_no_{p}') or '').strip(),
                (form.get(f'item_remarks_{p}') or '/').strip(),
            ]
            num_cols = (4, 6, 8, 9, 11)

        for j, v in enumerate(vals, 1):
            c = ws.cell(r, j)
            c.value = v
            c.font = openpyxl.styles.Font(size=9)
            c.border = thin_border
            if j in num_cols:
                c.alignment = openpyxl.styles.Alignment(horizontal='right')
                if isinstance(v, float):
                    c.number_format = '0.00'
            elif j == (10 if is_tile else 5):
                c.alignment = openpyxl.styles.Alignment(horizontal='center')
            else:
                c.alignment = openpyxl.styles.Alignment(horizontal='center' if j == 1 else 'left')
        ws.row_dimensions[r].height = 24

    last_item_row = row + n - 1
    row = last_item_row + 1

    # Total CTNS
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    ws.cell(row, 10).value = 'Total CTNS:'
    ws.cell(row, 10).font = openpyxl.styles.Font(bold=True, size=11)
    ws.cell(row, 12).value = round(total_ctns)
    ws.cell(row, 12).font = openpyxl.styles.Font(bold=True, size=12)
    ws.cell(row, 12).alignment = openpyxl.styles.Alignment(horizontal='right')
    for c in range(1, 15):
        ws.cell(row, c).border = thin_border
    ws.row_dimensions[row].height = 19
    row += 1

    # Total G.W.
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    ws.cell(row, 10).value = 'Total G.W. (kg):'
    ws.cell(row, 10).font = openpyxl.styles.Font(bold=True, size=11)
    ws.cell(row, 12).value = round(total_gw, 2)
    ws.cell(row, 12).font = openpyxl.styles.Font(bold=True, size=12)
    ws.cell(row, 12).number_format = '0.00'
    ws.cell(row, 12).alignment = openpyxl.styles.Alignment(horizontal='right')
    for c in range(1, 15):
        ws.cell(row, c).border = thin_border
    ws.row_dimensions[row].height = 19
    row += 1

    # Total N.W.
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    ws.cell(row, 10).value = 'Total N.W. (kg):'
    ws.cell(row, 10).font = openpyxl.styles.Font(bold=True, size=11)
    ws.cell(row, 12).value = round(total_nw, 2)
    ws.cell(row, 12).font = openpyxl.styles.Font(bold=True, size=12)
    ws.cell(row, 12).number_format = '0.00'
    ws.cell(row, 12).alignment = openpyxl.styles.Alignment(horizontal='right')
    for c in range(1, 15):
        ws.cell(row, c).border = thin_border
    ws.row_dimensions[row].height = 19
    row += 1

    # Total CBM (sanitary only)
    if not is_tile:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
        ws.cell(row, 10).value = 'Total CBM (m³):'
        ws.cell(row, 10).font = openpyxl.styles.Font(bold=True, size=11)
        ws.cell(row, 12).value = round(total_cbm, 3)
        ws.cell(row, 12).font = openpyxl.styles.Font(bold=True, size=12)
        ws.cell(row, 12).number_format = '0.000'
        ws.cell(row, 12).alignment = openpyxl.styles.Alignment(horizontal='right')
        for c in range(1, 15):
            ws.cell(row, c).border = thin_border
        ws.row_dimensions[row].height = 19
        row += 1

    # Stamp
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    ws.cell(row, 1).value = 'Company Stamp  公司印章:' if not is_ar else 'Company Stamp  ختم الشركة:'
    ws.cell(row, 1).font = openpyxl.styles.Font(size=10, color='555555')
    for c in range(1, 15):
        ws.cell(row, c).border = thin_border
    ws.row_dimensions[row].height = 30
    row += 1

    # Logo insertion (A1:B2 area)
    selected_logo = form.get('selected_logo_name', '').strip()
    if selected_logo:
        logo_path = os.path.join(LOGO_DIR, selected_logo)
        if os.path.exists(logo_path):
            with open(logo_path, 'rb') as lf:
                excel_insert_image(ws, lf, 'A1', w=150, h=70)

    # Print area
    last_row = row - 1
    ws.print_area = f'A1:N{last_row}'

    # Output
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    inv_no   = (form.get('invoice_no') or 'PL').strip()
    customer = (form.get('customer_name') or '').strip()
    suffix   = '_AR' if is_ar else ''
    filename = f"PL_{inv_no}_{customer}{suffix}.xlsx".replace(' ', '_').replace('/', '-')

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    d = date.today()
    today     = f"{d.year}.{d.month}.{d.day}"
    today_iso = d.isoformat()
    logos = logo_list_with_names()
    return render_template('index.html', today=today, today_iso=today_iso, logos=logos,
                           active_tab='pi-maker')


@app.route('/pl-maker')
def pl_maker():
    d = date.today()
    today     = f"{d.year}.{d.month}.{d.day}"
    today_iso = d.isoformat()
    logos = logo_list_with_names()
    return render_template('pl_maker.html', today=today, today_iso=today_iso, logos=logos,
                           active_tab='pl-maker')


@app.route('/ci-maker')
def ci_maker():
    d = date.today()
    today     = f"{d.year}.{d.month}.{d.day}"
    today_iso = d.isoformat()
    logos = logo_list_with_names()
    return render_template('ci_maker.html', today=today, today_iso=today_iso, logos=logos,
                           active_tab='ci-maker')


@app.route('/quotation-maker')
def quotation_maker():
    return render_template('placeholder.html',
                           title='Quotation Maker — 报价单制作',
                           active_tab='quotation-maker')


@app.route('/jincheng-invoice')
def jincheng_invoice():
    return render_template('placeholder.html',
                           title='JINCHENG Invoice — 金城含税发票申请',
                           active_tab='jincheng-invoice')


@app.route('/logos')
def api_logos():
    return jsonify(logo_list_with_names())


@app.route('/upload-logo', methods=['POST'])
def upload_logo():
    f = request.files.get('logo_file')
    if not f or not f.filename:
        return jsonify({'error': 'No file provided'}), 400
    filename = secure_filename(f.filename)
    if not filename or os.path.splitext(filename)[1].lower() not in LOGO_EXTS:
        return jsonify({'error': 'Invalid file type'}), 400
    f.save(os.path.join(LOGO_DIR, filename))
    brand_name = request.form.get('brand_name', '').strip()
    if brand_name:
        names = load_logo_names()
        names[filename] = brand_name
        save_logo_names(names)
    display_name = brand_name or os.path.splitext(filename)[0].upper()
    return jsonify({'filename': filename, 'display_name': display_name})


@app.route('/delete-logo', methods=['POST'])
def delete_logo():
    filename = (request.form.get('filename') or '').strip()
    if not filename:
        return jsonify({'error': 'Filename required'}), 400
    if '/' in filename or '\\' in filename:
        return jsonify({'error': 'Invalid filename'}), 400
    path = os.path.join(LOGO_DIR, filename)
    if os.path.isfile(path):
        os.remove(path)
    names = load_logo_names()
    if filename in names:
        del names[filename]
        save_logo_names(names)
    return jsonify({'success': True})


@app.route('/logo/<path:filename>')
def serve_logo(filename):
    return send_from_directory(LOGO_DIR, filename)


@app.route('/customers')
def api_customers():
    return jsonify(load_customers())


@app.route('/save-customer', methods=['POST'])
def api_save_customer():
    data = request.get_json()
    if not data or not data.get('name', '').strip():
        return jsonify({'error': 'Customer name required'}), 400
    customers = load_customers()
    customer = {
        'id':      str(uuid.uuid4()),
        'name':    data.get('name',    '').strip(),
        'country': data.get('country', '').strip(),
        'address': data.get('address', '').strip(),
        'email':   data.get('email',   '').strip(),
        'tel':     data.get('tel',     '').strip(),
    }
    customers.append(customer)
    save_customers_file(customers)
    return jsonify(customer)


@app.route('/delete-customer/<cid>', methods=['DELETE'])
def api_delete_customer(cid):
    customers = [c for c in load_customers() if c.get('id') != cid]
    save_customers_file(customers)
    return jsonify({'ok': True})


@app.route('/products')
def api_products():
    products = load_products()
    project = request.args.get('project', '').strip()
    if project:
        products = [p for p in products if p.get('project', '') == project]
    return jsonify(products)


@app.route('/save-product', methods=['POST'])
def api_save_product():
    data = request.get_json() or {}
    if not data.get('desc', '').strip():
        return jsonify({'error': 'Product description required'}), 400
    products = load_products()
    product = {
        'id':          str(uuid.uuid4()),
        'project':     data.get('project',     '').strip(),
        'desc':        data.get('desc',        '').strip(),
        'model':       data.get('model',       '').strip(),
        'color':       data.get('color',       '').strip(),
        'size':        data.get('size',        '').strip(),
        'dia_shape':   data.get('dia_shape',   '').strip(),
        'lumin':       data.get('lumin',       '').strip(),
        'voltage':     data.get('voltage',     '').strip(),
        'power':       data.get('power',       '').strip(),
        'color_temp':  data.get('color_temp',  '').strip(),
        'material':    data.get('material',    '').strip(),
        'tile_type':   data.get('tile_type',   '').strip(),
        'thickness':   data.get('thickness',   '').strip(),
        'brand':       data.get('brand',       '').strip(),
        'm2_per_ctn':  data.get('m2_per_ctn',  '').strip(),
        'gw_per_ctn':  data.get('gw_per_ctn',  '').strip(),
        'nw_per_ctn':  data.get('nw_per_ctn',  '').strip(),
        'unit':        data.get('unit',        'Set').strip(),
        'price':       data.get('price',       '').strip(),
        'tax':         data.get('tax',         'excl').strip(),
        'packing':     data.get('packing',     'Carton').strip(),
        'moq':         data.get('moq',         '/').strip(),
        'lead_time':   data.get('lead_time',   '/').strip(),
        'remarks':     data.get('remarks',     '/').strip(),
    }
    products.append(product)
    save_products_file(products)
    return jsonify(product)


@app.route('/delete-product/<pid>', methods=['DELETE'])
def api_delete_product(pid):
    products = [p for p in load_products() if p.get('id') != pid]
    save_products_file(products)
    return jsonify({'ok': True})


@app.route('/bank-info')
def api_bank_info():
    banks = load_bank_info()
    project = request.args.get('project', '').strip()
    if project:
        banks = [b for b in banks if b.get('project') == project]
    return jsonify(banks)


@app.route('/save-bank-info', methods=['POST'])
def api_save_bank_info():
    data = request.get_json() or {}
    project = data.get('project', '').strip()
    if not project:
        return jsonify({'error': 'Project name required'}), 400
    banks = load_bank_info()
    existing = next((b for b in banks if b.get('project') == project), None)
    entry = {
        'id':           existing['id'] if existing else str(uuid.uuid4()),
        'project':      project,
        'project_en':   data.get('project_en', '').strip(),
        'beneficiary':  data.get('beneficiary', '').strip(),
        'bank':         data.get('bank', '').strip(),
        'swift':        data.get('swift', '').strip(),
        'account':      data.get('account', '').strip(),
        'iban':         data.get('iban', '').strip(),
        'bank_address': data.get('bank_address', '').strip(),
    }
    if existing:
        for i, b in enumerate(banks):
            if b.get('project') == project:
                banks[i] = entry
                break
    else:
        banks.append(entry)
    save_bank_info_file(banks)
    return jsonify(entry)


@app.route('/delete-bank-info/<bid>', methods=['DELETE'])
def api_delete_bank_info(bid):
    banks = [b for b in load_bank_info() if b.get('id') != bid]
    save_bank_info_file(banks)
    return jsonify({'ok': True})


@app.route('/translate', methods=['POST'])
def api_translate():
    data = request.get_json() or {}
    text = data.get('text', '').strip()
    if not text:
        return jsonify({'error': 'No text provided'}), 400
    try:
        params = urllib.parse.urlencode({'q': text, 'langpair': 'zh|en'})
        url = f'https://api.mymemory.translated.net/get?{params}'
        req = urllib.request.Request(url, headers={'User-Agent': 'PImaker/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode())
        translated = result.get('responseData', {}).get('translatedText', '')
        if translated:
            return jsonify({'translated': translated})
        return jsonify({'error': 'No translation returned'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/preview', methods=['POST'])
def preview():
    form = request.form
    files = request.files
    ctx = build_preview_context(form, files)
    is_ar = (form.get('lang', 'zh') == 'ar')
    is_tile = (form.get('project', '').strip() == '瓷砖')
    if is_tile:
        template = 'ceramic_pi_preview.html'
    elif is_ar:
        template = 'invoice_preview_ar.html'
    else:
        template = 'invoice_preview.html'
    return render_template(template, **ctx)


@app.route('/generate', methods=['POST'])
def generate():
    try:
        lang = request.form.get('lang', 'zh')
        return fill_template(request.form, request.files, lang=lang)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ── CI Maker routes ─────────────────────────────────────────────────────────

@app.route('/ci/generate', methods=['POST'])
def ci_generate():
    try:
        lang = request.form.get('lang', 'zh')
        return fill_template(request.form, request.files, lang=lang, mode='ci')
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/ci/preview', methods=['POST'])
def ci_preview():
    form = request.form
    files = request.files
    ctx = build_preview_context(form, files, mode='ci')
    is_ar = (form.get('lang', 'zh') == 'ar')
    is_tile = (form.get('project', '').strip() == '瓷砖')
    if is_tile:
        template = 'ceramic_pi_preview.html'
    elif is_ar:
        template = 'invoice_preview_ar.html'
    else:
        template = 'invoice_preview.html'
    return render_template(template, **ctx)


# ── PL Maker routes ─────────────────────────────────────────────────────────

@app.route('/pl/generate', methods=['POST'])
def pl_generate():
    try:
        lang = request.form.get('lang', 'zh')
        return fill_pl_template(request.form, request.files, lang=lang)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/pl/preview', methods=['POST'])
def pl_preview():
    form = request.form
    files = request.files
    ctx = build_pl_preview_context(form, files)
    is_ar = (form.get('lang', 'zh') == 'ar')
    template = 'pl_preview_ar.html' if is_ar else 'pl_preview.html'
    return render_template(template, **ctx)


if __name__ == '__main__':
    app.run(debug=True, port=5000)
