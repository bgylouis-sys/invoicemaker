"""Build the Ceramic Tile Proforma Invoice Excel template from scratch.

Run once to (re)generate `约旦瓷砖PI模板.xlsx`.
Layout: portrait A4, 14 cols, rows dynamic, fit-to-one-page.
Matches the existing PI template structure but with tile-specific columns.
"""
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

OUTFILE = '约旦瓷砖PI模板.xlsx'

# Borders
thin   = Side(style='thin',   color='000000')
medium = Side(style='medium', color='000000')
thick  = Side(style='thick',  color='000000')

BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_MED   = Border(left=medium, right=medium, top=medium, bottom=medium)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)

# Fonts
TITLE_FONT = Font(name='Arial', size=28, bold=True, color='000000')
HEAD_FONT  = Font(name='Arial', size=13, bold=True, color='000000')
LBL_FONT   = Font(name='Arial', size=11, bold=False, color='000000')
VAL_FONT   = Font(name='Arial', size=12, bold=True, color='000000')
ITEM_HFONT = Font(name='Arial', size=9,  bold=True, color='000000')
ITEM_FONT  = Font(name='Arial', size=9,  color='000000')
GRAND_FONT = Font(name='Arial', size=15, bold=True, color='000000')
FIX_FONT   = Font(name='Arial', size=9,  color='000000')
LOGO_FONT  = Font(name='Arial', size=11, italic=True, color='888888')
SIGN_FONT  = Font(name='Arial', size=10, color='333333')

# Alignments
CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_W   = Alignment(horizontal='left',   vertical='center', wrap_text=True, indent=1)
RIGHT_W  = Alignment(horizontal='right',  vertical='center', wrap_text=True, indent=1)
TOP_L    = Alignment(horizontal='left',   vertical='top',    wrap_text=True, indent=1)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proforma Invoice'

    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4,
                                   header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True

    # Column widths
    widths = {'A':4, 'B':18, 'C':7, 'D':8, 'E':6, 'F':9,
              'G':8, 'H':8, 'I':8, 'J':6, 'K':8,
              'L':10, 'M':10, 'N':10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Helpers
    def merge(rng):
        ws.merge_cells(rng)

    def border_range(rng, b=BORDER):
        for row in ws[rng]:
            for cell in row:
                cell.border = b

    def setc(coord, value=None, font=None, align=None, b=BORDER):
        c = ws[coord]
        if value is not None:
            c.value = value
        if font:  c.font = font
        if align: c.alignment = align
        if b:     c.border = b

    def section_bar(rng, text):
        merge(rng)
        first = rng.split(':')[0]
        setc(first, text, HEAD_FONT, CENTER, b=BORDER_MED)
        border_range(rng, BORDER_MED)

    def lv(label_rng, val_rng, label_text, val=None):
        merge(label_rng)
        merge(val_rng)
        setc(label_rng.split(':')[0], label_text, LBL_FONT, LEFT_W)
        setc(val_rng.split(':')[0],   val,        VAL_FONT, LEFT_W)
        border_range(label_rng)
        border_range(val_rng)

    # Row heights
    heights = {
        1:28, 2:28, 3:18,
        4:18, 5:18, 6:18, 7:18, 8:18, 9:18,
        10:18, 11:18, 12:18, 13:18, 14:18, 15:18,
        16:28,
        17:32, 18:32, 19:32, 20:32, 21:32,
        22:18, 23:18, 24:26, 25:18,
        26:15, 27:15, 28:15, 29:15, 30:15,
        31:18,
        32:18, 33:18, 34:18, 35:18, 36:30,
    }
    for r, h in heights.items():
        ws.row_dimensions[r].height = h

    # ═══ Rows 1-2: Logo + Title ═══
    merge('A1:D2')
    setc('A1', '[ Insert Company Logo ]', LOGO_FONT, CENTER)
    merge('E1:N2')
    setc('E1', 'PROFORMA INVOICE\n形式发票', TITLE_FONT, CENTER)

    # ═══ Row 3: Section bars ═══
    section_bar('A3:G3', 'FROM / SUPPLIER  供货商')
    section_bar('H3:N3', 'INVOICE INFO  发票信息')

    # ═══ Rows 4-8: Supplier + Invoice Info ═══
    lv('A4:B4', 'C4:G4', 'Company Name:')
    lv('H4:I4', 'J4:N4', 'Invoice No.:')
    lv('A5:B5', 'C5:G5', 'Address:')
    lv('H5:I5', 'J5:N5', 'Date:')
    lv('A6:B6', 'C6:G6', 'City / Country:')
    lv('H6:I6', 'J6:N6', 'Currency:')
    lv('A7:B7', 'C7:G7', 'Tel:')
    lv('H7:I7', 'J7:N7', 'Incoterms:')
    lv('A8:B8', 'C8:G8', 'Email:')
    lv('H8:I8', 'J8:N8', 'Payment Terms:')

    # ═══ Row 9: Customer bars ═══
    section_bar('A9:G9', 'TO / CUSTOMER  客户')
    section_bar('H9:N9', 'QUOTATION DETAILS  报价详情')

    # ═══ Rows 10-14: Customer + Quotation ═══
    lv('A10:B10', 'C10:G10', 'Customer Name:')
    lv('H10:I10', 'J10:N10', 'Contact Person:')
    lv('A11:B11', 'C11:G11', 'Customer Address:')
    lv('H11:I11', 'J11:N11', 'Project / Inquiry Ref.:')
    lv('A12:B12', 'C12:G12', 'Country:')
    lv('H12:I12', 'J12:N12', 'Lead Time:')
    lv('A13:B13', 'C13:G13', 'Customer Tel:')
    lv('H13:I13', 'J13:N13', 'Quotation Validity:')
    lv('A14:B14', 'C14:G14', 'MOQ:')
    lv('H14:I14', 'J14:N14', 'Shipping (Port→Dest):')

    # ═══ Row 15: ITEM LIST bar ═══
    section_bar('A15:N15', 'ITEM LIST  商品明细 — 瓷砖 / Ceramic Tiles')

    # ═══ Row 16: Item table headers ═══
    headers = [
        'Item\nNo.', 'ITEM NO.', 'TYPE', 'SIZE', 'THK\n(mm)', 'BRAND',
        'M²/CTN', 'GW/CTN\n(kg)', 'NW/CTN\n(kg)', 'Unit', 'Qty\n(m²)',
        'Unit Price\nUSD/m²', 'Amount\n(USD)', 'Remarks',
    ]
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        c = ws[f'{col}16']
        c.value = h
        c.font = ITEM_HFONT
        c.alignment = CENTER
        c.border = BORDER_MED

    # ═══ Rows 17-21: Item data slots (5 rows) ═══
    for r in range(17, 22):
        for i in range(1, 15):
            col = get_column_letter(i)
            align = LEFT_W if i in (14,) else CENTER
            setc(f'{col}{r}', None, ITEM_FONT, align)
        ws[f'L{r}'].number_format = '#,##0.00'
        ws[f'M{r}'].number_format = '#,##0.00'
        ws[f'M{r}'].value = f'=K{r}*L{r}'

    # ═══ Row 22: Subtotal ═══
    merge('A22:G22')
    setc('A22', None, ITEM_FONT, LEFT_W)
    border_range('A22:G22')
    lv('H22:I22', 'J22:N22', 'Subtotal:')
    ws['J22'].number_format = '#,##0.00'
    ws['J22'].alignment = RIGHT_W
    ws['J22'].value = '=SUM(M17:M21)'

    # ═══ Row 23: Freight ═══
    merge('A23:G23')
    setc('A23', None, ITEM_FONT, LEFT_W)
    border_range('A23:G23')
    lv('H23:I23', 'J23:N23', 'Freight / Additional Charges:')
    ws['J23'].number_format = '#,##0.00'
    ws['J23'].alignment = RIGHT_W

    # ═══ Row 24: Grand Total ═══
    merge('A24:G24')
    setc('A24', None, ITEM_FONT, LEFT_W, b=BORDER_THICK)
    border_range('A24:G24', BORDER_THICK)
    merge('H24:I24')
    setc('H24', 'GRAND TOTAL:', GRAND_FONT, CENTER, b=BORDER_THICK)
    border_range('H24:I24', BORDER_THICK)
    merge('J24:M24')
    setc('J24', '=J22+J23', GRAND_FONT, RIGHT_W, b=BORDER_THICK)
    ws['J24'].number_format = '#,##0.00'
    border_range('J24:M24', BORDER_THICK)
    setc('N24', None, GRAND_FONT, CENTER, b=BORDER_THICK)

    # ═══ Row 25: REMARKS bar ═══
    section_bar('A25:N25', 'REMARKS  备注')

    # ═══ Rows 26-30: Fixed remarks ═══
    remarks = [
        '1.  The above prices are based on the stated quantity and specifications only.',
        '2.  Prices may be adjusted if there is any change in quantity, specification, packaging, or surface finish.',
        '3.  Standard export packing is included unless otherwise specified.',
        '4.  This quotation is valid for [  ] days from the date of issue.',
        '5.  Final supply terms shall be subject to order confirmation.',
    ]
    for i, text in enumerate(remarks):
        r = 26 + i
        merge(f'A{r}:N{r}')
        setc(f'A{r}', text, FIX_FONT, LEFT_W)
        border_range(f'A{r}:N{r}')

    # ═══ Row 31: TRANSFER DETAILS bar ═══
    section_bar('A31:N31', 'TRANSFER DETAILS  转账信息')

    # ═══ Rows 32-35: Bank info ═══
    lv('A32:B32', 'C32:G32', 'Beneficiary:')
    lv('H32:I32', 'J32:N32', 'Bank:')
    lv('A33:B33', 'C33:G33', 'SWIFT:')
    lv('H33:I33', 'J33:N33', 'Account:')
    lv('A34:B34', 'C34:N34', 'IBAN:')
    lv('A35:B35', 'C35:N35', 'Bank Address:')

    # ═══ Row 36: Stamp ═══
    merge('A36:N36')
    setc('A36', 'Company Stamp  公司印章:', SIGN_FONT, TOP_L)

    ws.print_area = 'A1:N36'
    wb.save(OUTFILE)
    print(f'Saved: {OUTFILE}')


if __name__ == '__main__':
    build()
