"""Shared Excel utilities for PIMaker templates.

Extracted from app.py fill_template() so CIPL and ceramic PI can reuse the same
merged-cell-safe row ops, logo insertion, date formatting, and bank filling.
"""

import io
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


# ── Style helpers ──────────────────────────────────────────────────────────────

def copy_row_style(ws, src_row, dst_row, max_col=14):
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst.font      = copy(src.font)
            dst.fill      = copy(src.fill)
            dst.border    = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format


# ── Image / Logo ───────────────────────────────────────────────────────────────

def insert_image(ws, file_obj, anchor, w=100, h=75):
    try:
        raw = file_obj.read()
        pil = PILImage.open(io.BytesIO(raw))
        pil.thumbnail((w, h), PILImage.LANCZOS)
        buf = io.BytesIO()
        fmt = (pil.format or 'PNG').upper()
        if fmt not in ('PNG', 'JPEG', 'GIF', 'BMP'):
            fmt = 'PNG'
        if fmt == 'JPEG' and pil.mode == 'RGBA':
            pil = pil.convert('RGB')
        pil.save(buf, format=fmt)
        buf.seek(0)
        xl_img = XLImage(buf)
        xl_img.anchor = anchor
        ws.add_image(xl_img)
    except Exception as e:
        print(f'[image] {anchor}: {e}')


# ── Date formatting ────────────────────────────────────────────────────────────

def format_invoice_date(date_raw):
    if date_raw and '-' in date_raw:
        try:
            d_obj = datetime.strptime(date_raw, '%Y-%m-%d')
            return f"{d_obj.year}.{d_obj.month}.{d_obj.day}"
        except ValueError:
            return date_raw
    return date_raw


# ── Merged-cell-safe row shift ─────────────────────────────────────────────────
#
# openpyxl's insert_rows / delete_rows corrupt merged-cell references,
# so we save all merges below the affected area, remove them, perform the
# row operation, then re-merge at the correct shifted positions.


def _merges_below(ws, start_row):
    """Collect merged ranges whose min_row >= start_row."""
    out = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row >= start_row:
            out.append((rng.min_row, rng.max_row, rng.min_col, rng.max_col))
    return out


def _remove_merges_below(ws, start_row):
    """Remove merged ranges at or below start_row (direct list removal)."""
    to_del = [rng for rng in ws.merged_cells.ranges if rng.min_row >= start_row]
    for rng in to_del:
        ws.merged_cells.ranges.remove(rng)


def _reapply_merges(ws, saved, row_shift, items_start):
    """Re-merge saved ranges after applying a row shift, skipping those
    that would land above items_start."""
    for min_r, max_r, min_c, max_c in saved:
        nr_min = min_r + row_shift
        nr_max = max_r + row_shift
        if nr_min >= items_start:
            ws.merge_cells(start_row=nr_min, start_column=min_c,
                           end_row=nr_max, end_column=max_c)


def shift_item_rows(ws, items_start, items_end, n_items):
    """Insert/delete item rows in a merged-cell-safe way.

    Args:
        ws:          The worksheet.
        items_start: First data row (1-based).
        items_end:   Last template data row (1-based).
        n_items:     Actual number of items to show.

    Returns:
        (offset, deleted, shift): row_count_added, row_count_deleted, net_shift
    """
    slots = items_end - items_start + 1
    offset = 0
    deleted = 0

    if n_items > slots:
        extra = n_items - slots
        offset = extra
        insert_at = items_end + 1

        saved = _merges_below(ws, insert_at)
        _remove_merges_below(ws, insert_at)

        ws.insert_rows(insert_at, extra)

        # Remove phantom ranges at the insertion point
        stale = [rng for rng in ws.merged_cells.ranges
                 if insert_at <= rng.min_row < insert_at + extra]
        for rng in stale:
            ws.merged_cells.ranges.remove(rng)

        _reapply_merges(ws, saved, extra, items_start)

        for i in range(extra):
            copy_row_style(ws, items_end, insert_at + i)

    if n_items < slots:
        deleted = slots - n_items
        delete_at = items_start + n_items

        saved = _merges_below(ws, delete_at)
        _remove_merges_below(ws, delete_at)

        ws.delete_rows(delete_at, deleted)

        # Remove phantom ranges at the deletion point
        stale = [rng for rng in ws.merged_cells.ranges
                 if delete_at <= rng.min_row < delete_at + deleted]
        for rng in stale:
            ws.merged_cells.ranges.remove(rng)

        _reapply_merges(ws, saved, -deleted, items_start)

    shift = offset - deleted
    return offset, deleted, shift


# ── Bank info fill ─────────────────────────────────────────────────────────────

def fill_bank_from_json(ws, base_row, bank_dict):
    """Fill bank info rows starting at base_row, unmerging first.

    Args:
        ws:        The worksheet.
        base_row:  First bank data row (e.g. 32 + shift).
        bank_dict: Dict with keys: beneficiary, bank, swift, account, iban, bank_address.
    """
    if not bank_dict:
        return

    br = base_row
    # Unmerge bank rows first (delete_rows may have corrupted merged ranges)
    for rr in range(br, br + 4):
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= rr <= merged_range.max_row:
                ws.unmerge_cells(str(merged_range))

    ws.cell(br,     3).value = bank_dict.get('beneficiary', '')
    ws.cell(br,    10).value = bank_dict.get('bank', '')
    ws.cell(br + 1, 3).value = bank_dict.get('swift', '')
    acct = bank_dict.get('account', '')
    curr = bank_dict.get('currency', '')
    ws.cell(br + 1,10).value = f'{acct} ({curr})' if curr else acct
    ws.cell(br + 2, 3).value = bank_dict.get('iban', '')
    ws.cell(br + 3, 3).value = bank_dict.get('bank_address', '')
