"""Build the Proforma Invoice Excel template from scratch.

Run once to (re)generate `约旦卫浴Proforma Invoice模板.xlsx`.
Layout: portrait A4, 14 cols, 38 rows, fit-to-one-page.
Pure black & white (no fills), bold borders, large fonts.
Matches templates/invoice_preview.html exactly.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

OUTFILE = '约旦卫浴Proforma Invoice模板.xlsx'

# Borders
thin   = Side(style='thin',   color='000000')
medium = Side(style='medium', color='000000')
thick  = Side(style='thick',  color='000000')

BORDER          = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_BAR      = Border(left=medium, right=medium, top=medium, bottom=medium)
BORDER_GRAND    = Border(left=thick, right=thick, top=thick, bottom=thick)

# Fonts (large for print clarity)
TITLE_FONT  = Font(name='Arial', size=28, bold=True, color='000000')
HEAD_FONT   = Font(name='Arial', size=13, bold=True, color='000000')
LBL_FONT    = Font(name='Arial', size=11, bold=False, color='000000')
VAL_FONT    = Font(name='Arial', size=12, bold=True,  color='000000')
ITEM_HFONT  = Font(name='Arial', size=10, bold=True,  color='000000')
ITEM_FONT   = Font(name='Arial', size=10, color='000000')
GRAND_FONT  = Font(name='Arial', size=15, bold=True,  color='000000')
FIX_FONT    = Font(name='Arial', size=10, color='000000')
LOGO_FONT   = Font(name='Arial', size=11, italic=True, color='888888')
SIGN_FONT = Font(name='Arial', size=10, color='333333')

# Alignments
CENTER   = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT     = Alignment(horizontal='left',   vertical='center', wrap_text=True, indent=1)
RIGHT    = Alignment(horizontal='right',  vertical='center', wrap_text=True, indent=1)
TOP_LEFT = Alignment(horizontal='left',   vertical='top',    wrap_text=True, indent=1)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Proforma Invoice'

    # Portrait A4, fit to one page
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.4, bottom=0.4,
                                  header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True

    # Column widths
    widths = {
        'A': 4,    'B': 13,   'C': 9,    'D': 12,
        'E': 9,    'F': 11,   'G': 5,    'H': 6,
        'I': 10,   'J': 11,   'K': 8,    'L': 5,
        'M': 7,    'N': 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Helpers
    def merge(rng):
        ws.merge_cells(rng)

    def border_range(rng, b=BORDER):
        for row in ws[rng]:
            for cell in row:
                cell.border = b

    def setc(coord, value=None, font=None, align=None, b=BORDER):
        c = ws[coord]
        if value is not None:
            c.value = value
        if font:  c.font = font
        if align: c.alignment = align
        if b:     c.border = b

    def section_bar(rng, text):
        merge(rng)
        first = rng.split(':')[0]
        setc(first, text, HEAD_FONT, CENTER, b=BORDER_BAR)
        border_range(rng, BORDER_BAR)

    def lv(label_rng, val_rng, label_text, val=None):
        merge(label_rng)
        merge(val_rng)
        setc(label_rng.split(':')[0], label_text, LBL_FONT, LEFT)
        setc(val_rng.split(':')[0],   val,        VAL_FONT, LEFT)
        border_range(label_rng)
        border_range(val_rng)

    # Row heights (38 rows, sized to fit A4 portrait with larger fonts)
    heights = {
        1: 28, 2: 28,
        3: 18,
        4: 18, 5: 18, 6: 18, 7: 18, 8: 18,
        9: 18,
        10: 18, 11: 18, 12: 18, 13: 18, 14: 18,
        15: 18,
        16: 24,
        17: 36, 18: 36, 19: 36, 20: 36, 21: 36,
        22: 18, 23: 18,
        24: 26,
        25: 18,
        26: 15, 27: 15, 28: 15, 29: 15, 30: 15,
        31: 18,
        32: 18, 33: 18, 34: 18, 35: 18,
        36: 30,
    }
    for r, h in heights.items():
        ws.row_dimensions[r].height = h

    # ═══ ROW 1-2: Logo + Title ═══
    merge('A1:D2')
    setc('A1', '[ Insert Company Logo ]', LOGO_FONT, CENTER)
    merge('E1:N2')
    setc('E1', 'PROFORMA INVOICE', TITLE_FONT, CENTER)

    # ═══ ROW 3: Section bars ═══
    section_bar('A3:F3', 'FROM / SUPPLIER  供货商')
    section_bar('G3:N3', 'INVOICE INFO  发票信息')

    # ═══ ROWS 4-8: Supplier (left) + Invoice/Terms (right) ═══
    lv('A4:B4', 'C4:F4', 'Company Name:')
    lv('G4:I4', 'J4:N4', 'Invoice No.:')
    lv('A5:B5', 'C5:F5', 'Address:')
    lv('G5:I5', 'J5:N5', 'Date:')
    lv('A6:B6', 'C6:F6', 'City / Country:')
    lv('G6:I6', 'J6:N6', 'Currency:')
    lv('A7:B7', 'C7:F7', 'Tel:')
    lv('G7:I7', 'J7:N7', 'Incoterms:')
    lv('A8:B8', 'C8:F8', 'Email:')
    lv('G8:I8', 'J8:N8', 'Payment Terms:')

    # ═══ ROW 9: Customer + Quotation Details bars ═══
    section_bar('A9:F9', 'TO / CUSTOMER  客户')
    section_bar('G9:N9', 'QUOTATION DETAILS  报价详情')

    # ═══ ROWS 10-14: Customer (left) + Quotation Details (right) ═══
    lv('A10:B10', 'C10:F10', 'Customer Name:')
    lv('G10:I10', 'J10:N10', 'Contact Person:')
    lv('A11:B11', 'C11:F11', 'Customer Address:')
    lv('G11:I11', 'J11:N11', 'Project / Inquiry Ref.:')
    lv('A12:B12', 'C12:F12', 'Customer Email:')
    lv('G12:I12', 'J12:N12', 'Lead Time:')
    lv('A13:B13', 'C13:F13', 'Country:')
    lv('G13:I13', 'J13:N13', 'Quotation Validity:')
    lv('A14:B14', 'C14:F14', 'Customer Tel:')
    lv('G14:I14', 'J14:N14', 'Shipping (Port→Dest):')

    # ═══ ROW 15: ITEM LIST bar ═══
    section_bar('A15:N15', 'ITEM LIST  商品明细')

    # ═══ ROW 16: Item table headers ═══
    headers = ['Item\nNo.', 'Product Description', 'Model No.', None,
               'Color/\nFinish', 'Size', 'Unit', 'Qty', 'Unit Price',
               'Amount', None, None, None, 'Remarks']
    for i, h in enumerate(headers, start=1):
        col = get_column_letter(i)
        c = ws[f'{col}16']
        c.value = h
        c.font = ITEM_HFONT
        c.alignment = CENTER
        c.border = BORDER_BAR

    # ═══ ROWS 17-21: Item data rows (5 slots, may be deleted by app) ═══
    for r in range(17, 22):
        for i in range(1, 15):
            col = get_column_letter(i)
            align = LEFT if i in (2, 14) else CENTER
            setc(f'{col}{r}', None, ITEM_FONT, align)
        ws[f'I{r}'].number_format = '#,##0.00'
        ws[f'J{r}'].number_format = '#,##0.00'
        ws[f'J{r}'].value = f'=H{r}*I{r}'

    # ═══ ROW 22: Subtotal ═══
    merge('A22:F22')
    setc('A22', None, ITEM_FONT, LEFT)
    border_range('A22:F22')
    lv('G22:I22', 'J22:N22', 'Subtotal:')
    ws['J22'].number_format = '#,##0.00'
    ws['J22'].alignment = RIGHT
    ws['J22'].value = '=SUM(J17:J21)'

    # ═══ ROW 23: Freight ═══
    merge('A23:F23')
    setc('A23', None, ITEM_FONT, LEFT)
    border_range('A23:F23')
    lv('G23:I23', 'J23:N23', 'Freight / Additional Charges:')
    ws['J23'].number_format = '#,##0.00'
    ws['J23'].alignment = RIGHT

    # ═══ ROW 24: Grand Total ═══
    merge('A24:F24')
    setc('A24', None, ITEM_FONT, LEFT, b=BORDER_GRAND)
    border_range('A24:F24', BORDER_GRAND)
    merge('G24:I24')
    setc('G24', 'GRAND TOTAL:', GRAND_FONT, CENTER, b=BORDER_GRAND)
    border_range('G24:I24', BORDER_GRAND)
    merge('J24:M24')
    setc('J24', '=J22+J23', GRAND_FONT, RIGHT, b=BORDER_GRAND)
    ws['J24'].number_format = '#,##0.00'
    border_range('J24:M24', BORDER_GRAND)
    setc('N24', None, GRAND_FONT, CENTER, b=BORDER_GRAND)

    # ═══ ROW 25: REMARKS bar ═══
    section_bar('A25:N25', 'REMARKS  备注')

    # ═══ ROWS 26-30: Fixed remarks ═══
    remarks = [
        '1.  The above prices are based on the stated quantity and specifications only.',
        '2.  Prices may be adjusted if there is any change in quantity, specification, packaging, or surface finish.',
        '3.  Standard export packing is included unless otherwise specified.',
        '4.  This quotation is valid for [  ] days from the date of issue.',
        '5.  Final supply terms shall be subject to order confirmation.',
    ]
    for i, text in enumerate(remarks):
        r = 26 + i
        merge(f'A{r}:N{r}')
        setc(f'A{r}', text, FIX_FONT, LEFT)
        border_range(f'A{r}:N{r}')

    # ═══ ROW 31: TRANSFER DETAILS bar ═══
    section_bar('A31:N31', 'TRANSFER DETAILS  转账信息')

    # ═══ ROWS 32-35: Bank info ═══
    lv('A32:B32', 'C32:G32', 'Beneficiary:', 'Jinsheng International Ceramics')
    lv('H32:I32', 'J32:N32', 'Bank:',        'Bank al Etihad')
    lv('A33:B33', 'C33:G33', 'SWIFT:',       'UBSIJOAX')
    lv('H33:I33', 'J33:N33', 'Account:',     '0250193398615101')
    lv('A34:B34', 'C34:N34', 'IBAN:',        'JO26 UBSI 1150 0002 5019 3398 6151 01')
    lv('A35:B35', 'C35:N35', 'Bank Address:','Bank al Etihad, Al Shmaisani, Headquarters, Amman, Jordan')

    # ═══ ROW 36: Stamp area ═══
    merge('A36:N36')
    setc('A36', 'Company Stamp  公司印章:', SIGN_FONT, TOP_LEFT)

    ws.print_area = 'A1:N36'
    wb.save(OUTFILE)
    print(f'Saved: {OUTFILE}')


if __name__ == '__main__':
    build()
